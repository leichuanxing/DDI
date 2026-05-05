import ipaddress
from io import BytesIO
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import Workbook, load_workbook
from .models import AddressSpace, IPAddress, IPAddressHistory, Subnet


class IPAMService:
    ALLOCATABLE_STATUSES = {'available', 'reserved'}
    CLEAR_FIELDS = {
        'hostname': '',
        'mac_address': '',
        'usage_type': '',
        'owner': '',
        'department': '',
        'description': '',
        'dns_record': None,
        'dhcp_reservation': None,
    }

    @staticmethod
    @transaction.atomic
    def generate_ips(subnet: Subnet):
        subnet.full_clean()
        subnet.save()
        created = 0
        skipped = 0
        for ip in subnet.network.hosts():
            ip_text = str(ip)
            if subnet.gateway and ip_text == str(subnet.gateway):
                skipped += 1
                continue
            _, was_created = IPAddress.objects.get_or_create(subnet=subnet, ip_address=ip_text, defaults={'status': 'available'})
            if was_created:
                created += 1
        return {'created': created, 'skipped': skipped, 'total': subnet.ip_addresses.count()}

    @staticmethod
    @transaction.atomic
    def set_status(ip_obj, status, user=None, action='status_change', **fields):
        old = ip_obj.status
        for key, value in fields.items():
            setattr(ip_obj, key, value)
        ip_obj.status = status
        ip_obj.full_clean()
        ip_obj.save()
        IPAddressHistory.objects.create(ip_address=ip_obj, action=action, old_status=old, new_status=status, operator=user, detail=fields)
        return ip_obj

    @classmethod
    @transaction.atomic
    def allocate(cls, ip_obj, user=None, **fields):
        if ip_obj.status not in cls.ALLOCATABLE_STATUSES:
            raise ValidationError({'status': f'当前状态为 {ip_obj.get_status_display()}，不能直接分配'})
        return cls.set_status(ip_obj, 'used', user, 'allocate', **fields)

    @classmethod
    @transaction.atomic
    def reserve(cls, ip_obj, user=None, **fields):
        if ip_obj.status not in {'available', 'used'}:
            raise ValidationError({'status': f'当前状态为 {ip_obj.get_status_display()}，不能设置预留'})
        return cls.set_status(ip_obj, 'reserved', user, 'reserve', **fields)

    @classmethod
    @transaction.atomic
    def release(cls, ip_obj, user=None):
        if ip_obj.status in {'dhcp_dynamic', 'dhcp_reserved'}:
            raise ValidationError({'status': 'DHCP 联动地址不能从 IPAM 手动释放，请先处理 DHCP 租约或保留地址'})
        return cls.set_status(ip_obj, 'available', user, 'release', **cls.CLEAR_FIELDS)

    @staticmethod
    def utilization():
        rows = []
        for subnet in Subnet.objects.select_related('address_space'):
            qs = subnet.ip_addresses.all()
            total = qs.count()
            available = qs.filter(status='available').count()
            used = qs.filter(status='used').count()
            reserved = qs.filter(status='reserved').count()
            dhcp_dynamic = qs.filter(status='dhcp_dynamic').count()
            dhcp_reserved = qs.filter(status='dhcp_reserved').count()
            used_total = total - available
            rate = round((used_total / total) * 100, 2) if total else 0
            rows.append({'subnet_id': subnet.id, 'cidr': subnet.cidr, 'address_space': subnet.address_space.name, 'total': total, 'available': available, 'used': used, 'reserved': reserved, 'dhcp_dynamic': dhcp_dynamic, 'dhcp_reserved': dhcp_reserved, 'utilization': rate, 'alert': 'red' if rate >= 90 else 'yellow' if rate >= 80 else 'normal'})
        return rows

    @classmethod
    def subnet_utilization(cls, subnet):
        return next((row for row in cls.utilization() if row['subnet_id'] == subnet.id), None)

    @staticmethod
    def export_subnets():
        wb = Workbook()
        ws = wb.active
        ws.title = 'subnets'
        ws.append(['地址空间编码', 'CIDR', '网关', '掩码', 'VLAN ID', 'VLAN 名称', '位置', '用途', '状态', '描述'])
        for subnet in Subnet.objects.select_related('address_space').order_by('address_space__code', 'cidr'):
            ws.append([
                subnet.address_space.code, subnet.cidr, subnet.gateway or '', subnet.netmask,
                subnet.vlan_id or '', subnet.vlan_name, subnet.location, subnet.usage_type,
                subnet.status, subnet.description,
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_ip_addresses():
        wb = Workbook()
        ws = wb.active
        ws.title = 'ip_addresses'
        ws.append(['网段', 'IP 地址', '主机名', 'MAC 地址', '状态', '用途', '使用人', '部门', '描述'])
        for ip_obj in IPAddress.objects.select_related('subnet').order_by('subnet__cidr', 'ip_address'):
            ws.append([
                ip_obj.subnet.cidr, str(ip_obj.ip_address), ip_obj.hostname, ip_obj.mac_address,
                ip_obj.status, ip_obj.usage_type, ip_obj.owner, ip_obj.department, ip_obj.description,
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    @transaction.atomic
    def import_subnets(file_obj):
        wb = load_workbook(file_obj)
        ws = wb.active
        created = 0
        updated = 0
        errors = []
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0] or not row[1]:
                continue
            try:
                address_space = AddressSpace.objects.get(code=str(row[0]).strip())
                cidr = str(ipaddress.ip_network(str(row[1]).strip(), strict=False))
                defaults = {
                    'gateway': str(row[2]).strip() if row[2] else None,
                    'netmask': str(row[3]).strip() if row[3] else '',
                    'vlan_id': int(row[4]) if row[4] not in (None, '') else None,
                    'vlan_name': str(row[5]).strip() if row[5] else '',
                    'location': str(row[6]).strip() if row[6] else '',
                    'usage_type': str(row[7]).strip() if row[7] else '',
                    'status': str(row[8]).strip() if row[8] else 'enabled',
                    'description': str(row[9]).strip() if len(row) > 9 and row[9] else '',
                }
                subnet, was_created = Subnet.objects.update_or_create(
                    address_space=address_space,
                    cidr=cidr,
                    defaults=defaults,
                )
                subnet.full_clean()
                subnet.save()
                created += 1 if was_created else 0
                updated += 0 if was_created else 1
            except Exception as exc:
                errors.append({'row': row_no, 'error': str(exc)})
        return {'created': created, 'updated': updated, 'errors': errors}

    @staticmethod
    @transaction.atomic
    def import_ip_addresses(file_obj, user=None):
        wb = load_workbook(file_obj)
        ws = wb.active
        updated = 0
        errors = []
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0] or not row[1]:
                continue
            try:
                subnet = Subnet.objects.get(cidr=str(row[0]).strip())
                ip_obj = IPAddress.objects.get(subnet=subnet, ip_address=str(row[1]).strip())
                fields = {
                    'hostname': str(row[2]).strip() if row[2] else '',
                    'mac_address': str(row[3]).strip().lower().replace('-', ':') if row[3] else '',
                    'usage_type': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                    'owner': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                    'department': str(row[7]).strip() if len(row) > 7 and row[7] else '',
                    'description': str(row[8]).strip() if len(row) > 8 and row[8] else '',
                }
                status = str(row[4]).strip() if len(row) > 4 and row[4] else ip_obj.status
                IPAMService.set_status(ip_obj, status, user, 'import_update', **fields)
                updated += 1
            except Exception as exc:
                errors.append({'row': row_no, 'error': str(exc)})
        return {'updated': updated, 'errors': errors}
